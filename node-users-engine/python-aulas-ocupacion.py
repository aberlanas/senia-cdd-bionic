#!/usr/bin/python3

import sys
import os
import random
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import openpyxl
import argparse

class Sesion:

        def __init__(self,dia_semana,sesion_orden, aula, docente, grupo,contenido):
                self.dia_semana = dia_semana
                self.sesion_orden = sesion_orden
                self.aula = aula
                self.docente = docente
                self.grupo = grupo
                self.contenido = contenido

        def print_sesion(self):
                print("DIA: "+self.dia_semana+ " MATERIA: "+self.contenido)

class Profesor:

        def __init__(self,documento,nombre,apellido1,apellido2,color):
                self.documento = documento
                self.nombre = nombre
                self.apellidos = [ apellido1,apellido2 ]
                self.color = color

class Grupo:

        def __init__(self,grupo,color):
                self.grupo=grupo
                self.color=color


horarios_file='/home/aberlanas/Descargas/imexalum.xml'
profes_file='/home/aberlanas/Descargas/imexalum.xml'

wb = Workbook()
ws = wb.active

# TODO mejorar esto como parametro
tree = ET.parse(horarios_file)
root = tree.getroot()

tree_profes = ET.parse(profes_file)
root_profes= tree_profes.getroot()

# ListadoProfes
listado_profes = []

for listadocentes in root_profes.findall("docentes"):
        for lista_docente in listadocentes.iter():
                for docenteaux in lista_docente:
                        random_color = ''.join([random.choice('0123456789ABCDEF') for j in range(6)])
                        docente = Profesor(docenteaux.attrib['documento'],docenteaux.attrib['nombre'],docenteaux.attrib['apellido1'],docenteaux.attrib['apellido2'],random_color)
                        listado_profes.append(docente)


def busca_docente(documento):
        docente_aux = Profesor("0000000A","Sin nombre","Sin apellido","Sin apellido","FFFFFF")
        for docente in listado_profes:
                if docente.documento == documento:
                        docente_aux = docente

        return docente_aux

# carga profesores

# Set de grupos con color
set_grupos=set()
set_grupos_color=set()

# Lista de Sesiones
lista_sesiones = []
set_aulas=set()

def busca_aulas_vacias():
    
    # Creamos un excel
    wb_vacia = Workbook()
    ws_vacia = wb_vacia.active
    
    # Aulas sesiones
    columna_vacia=1
    fila_vacia=1
    for dia in [ "L","M","X","J","V"]:
        for hora in ["1","2","3","5","6","7","9"]:
            print(" * [ Sesion: "+dia+":"+str(hora)+" ]")
            
            aulas_vacias=""
            for aula in set_aulas:
                libre=True
                
                for sesion in lista_sesiones:
                    #print(sesion.dia_semana+":"+sesion.sesion_orden+"->"+sesion.aula)
                    if (sesion.dia_semana==dia and sesion.sesion_orden==hora and sesion.aula==aula):
                        libre=False
                if (libre):
                    print(aula)
                    aulas_vacias+=aula+"\n"
            ws_vacia.cell(column=columna_vacia,row=fila_vacia,value=aulas_vacias)
            fila_vacia=fila_vacia+1
            
        fila_vacia=1
        columna_vacia=columna_vacia+1
    
    wb_vacia.save("patatas_vacias.xlsx")

def busca_grupo_color(grupo_nombre):
        auxcolor = "BBBBBB"
        for grupo in set_grupos_color:
                if (grupo.grupo == grupo_nombre):
                        auxcolor=grupo.color
        return auxcolor


def rellena_celda_sesion(columna,fila,sesion):
        aux_docente = busca_docente(sesion.docente)
        rgb_color = busca_grupo_color(sesion.grupo)
        aux_color = openpyxl.styles.colors.Color(rgb=rgb_color)
        aux_pattern = PatternFill(fill_type='solid', fgColor=aux_color)
        celda = ws.cell(column=columna,row=fila,value="["+sesion.grupo+"] "+sesion.contenido+" - "+aux_docente.nombre+" - "+aux_docente.apellidos[0])
        celda.fill=aux_pattern



for horario_grupo in root.findall("horarios_grupo"):
        for item in horario_grupo.iter():
                for campo in item:
                        if (str(campo.attrib['plantilla'])=="758772746"):
                                sesion = Sesion(campo.attrib['dia_semana'],campo.attrib['sesion_orden'],campo.attrib['aula'],campo.attrib['docente'],	campo.attrib['grupo'],campo.attrib['contenido'])
                                
                                
                                set_grupos.add(campo.attrib['grupo'])
                                lista_sesiones.append(sesion)
                                set_aulas.add(sesion.aula)
                                
# Asignamos ahora todos los colores a los grupos
for grup in set_grupos:
        random_color = ''.join([random.choice('0123456789ABCDEF') for j in range(6)])

        if(grup.startswith("1ESOA")):
                # RED 4422
                random_color = "1ABC9C"
                
        elif(grup.startswith("1ESOB")):
                # RED 4422
                random_color = "16A085"
                
        elif(grup.startswith("1ESOC")):
                # RED 4422
                random_color = "27AE60"
                
        elif(grup.startswith("1ESOD")):
                # RED 4422
                random_color = "2ECC71"
                
        elif(grup.startswith("1ESOE")):
                # RED 4422
                random_color = "2C0071"

        elif(grup.startswith("2ESOA")):
                # RED 4422
                random_color = "F1C40F"
                
        elif(grup.startswith("2ESOB")):
                # RED 4422
                random_color = "F39C12"
                
        elif(grup.startswith("2ESOC")):
                # RED 4422
                random_color = "E67E22"
        elif(grup.startswith("2ESOD")):
                # RED 4422
                random_color = "D35400"

        elif(grup.startswith("2ESOE")):
                # RED 4422
                random_color = "D35400"

        elif(grup.startswith("3ESOA")):
                # RED 4422
                random_color = "BDC3C7"
                
        elif(grup.startswith("3ESOB")):
                # RED 4422
                random_color = "95A5A6"
                
        elif(grup.startswith("3ESOC")):
                # RED 4422
                random_color = "7F8C8D"
        elif(grup.startswith("3ESOD")):
                # RED 4422
                random_color = "34495E"
        elif(grup.startswith("3ESOP")):
                # RED 4422
                random_color = "FABADA"
               
        elif(grup.startswith("4ESO")):
                # RED 8822
                random_color = random_color

        elif(grup.startswith("1BAHA") or grup.startswith("2BAHA")):
                # RED 4422
                random_color = random_color

        elif(grup.startswith("2C") or grup.startswith("1C")):
                # RED 4422
                random_color = random_color
                
                
        grupo = Grupo(grup,random_color)
        set_grupos_color.add(grupo)

for grupo_colorines in set_grupos_color:
        print(grupo_colorines.grupo+" -> "+grupo_colorines.color)
       
# Lista de sesiones por dia
sesiones_lunes = []
sesiones_martes = []
sesiones_miercoles = []
sesiones_jueves=[]
sesiones_viernes=[]


for sesion in lista_sesiones:
        if sesion.dia_semana=="L":
                sesiones_lunes.append(sesion)
        if sesion.dia_semana=="M":
                sesiones_martes.append(sesion)
        if sesion.dia_semana=="X":
                sesiones_miercoles.append(sesion)
        if sesion.dia_semana=="J":
                sesiones_jueves.append(sesion)
        if sesion.dia_semana=="V":
                sesiones_viernes.append(sesion)


fila = 2

for aula in set_aulas:
        columna=1
        color_top = openpyxl.styles.colors.Color(rgb="FFC0CB")
        color_top_fill = PatternFill(fill_type='solid', fgColor=color_top)
        ws.cell(column=columna,row=fila,value=aula)
        
        # LUNES
        for x in range(16):
                
                num_sesion =x+1
                columna=columna + 1
                celda_top = ws.cell(column=columna,row=1,value="LUNES : "+str(num_sesion))
                celda_top.fill=color_top_fill
                
                for sesion in sesiones_lunes:
                        if (str(sesion.sesion_orden) == str(num_sesion) and str(sesion.aula) == str(aula)):
                                rellena_celda_sesion(columna,fila,sesion)
        
        # MARTES
        for x in range(16):
                
                num_sesion=x+1
                columna=columna+1
                celda_top = ws.cell(column=columna,row=1,value="MARTES : "+str(num_sesion))
                celda_top.fill=color_top_fill

                for sesion in sesiones_martes:
                        if (str(sesion.sesion_orden) == str(num_sesion) and str(sesion.aula) == str(aula)):
                                rellena_celda_sesion(columna,fila,sesion)                                
        # MIERCOLES
        for x in range(16):
                
                num_sesion=x+1
                columna=columna+1
                celda_top = ws.cell(column=columna,row=1,value="MIERCOLES : "+str(num_sesion))
                celda_top.fill=color_top_fill

                for sesion in sesiones_miercoles:
                        if (str(sesion.sesion_orden) == str(num_sesion) and str(sesion.aula) == str(aula)):
                                rellena_celda_sesion(columna,fila,sesion)

        # JUEVES
        for x in range(16):
                
                num_sesion=x+1
                columna = columna+1
                celda_top = ws.cell(column=columna,row=1,value="JUEVES : "+str(num_sesion))
                celda_top.fill=color_top_fill

                for sesion in sesiones_jueves:                        
                        if (str(sesion.sesion_orden) == str(num_sesion) and str(sesion.aula) == str(aula)):
                                rellena_celda_sesion(columna,fila,sesion)
        # VIERNES
        for x in range(16):

                num_sesion=x+1
                columna = columna + 1
                celda_top = ws.cell(column=columna,row=1,value="VIERNES : "+str(num_sesion))
                celda_top.fill=color_top_fill

                for sesion in sesiones_viernes:
                        if (str(sesion.sesion_orden) == str(num_sesion) and str(sesion.aula) == str(aula)):
                                rellena_celda_sesion(columna,fila,sesion)
                                

        fila = fila+1


wb.save("patata.xlsx")


busca_aulas_vacias()


sys.exit(0)
