#!/usr/bin/python3
 # -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook

listFile = "/home/aberlanas/temp/GRUPOS_2021.xlsx"

wb = load_workbook(listFile)

for sheet_name in wb.sheetnames:

   sheet = wb[sheet_name]

   print( "" )
   print( "# Grup : " + sheet['A1'].value)
   print( "")

   print("|Nº| Nom |Cognoms|")
   print("|:-:|:---|:------|")
   
   numA = 1
   for row in sheet.iter_rows(min_row=4,max_row=24):
       nombre = None

       try:
           nombre = row[0].value
       except Exception as e:
           print(str(e))
           pass

       if nombre != None:
           #print (str(row[0].value))
           nombrel = nombre.split(',')
           
           try:
              nom = nombrel[1]
              ape = nombrel[0]
           except Exception as e:
              print(row[0].value)

           nomFinal = ""
           noml = nom.strip().split(' ')
           for auxn in noml:
               nomFinal += auxn.capitalize() + " "

           apelFinal = ""    
           apel = ape.strip().split(' ')

           for auxp in apel:
               auxf = auxp[0:3]

               sasteriscos = ""
               nasteriscos = len(auxp)-3
               for i in range(nasteriscos):
                   sasteriscos += "*"

               apelFinal+=auxf.capitalize()+sasteriscos+" "

           print ("| " + str(numA) + " | "+ nomFinal + "|" + apelFinal +"|")
           numA = numA + 1

   print(" ")
   print("---")
   print(" ")
   print("\\newpage ")
   print(" ")

